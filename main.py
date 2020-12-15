import telebot
import openpyxl
from tablework import WorkTable
from tokenBot import TOKEN


bot = telebot.TeleBot(TOKEN)
wt = WorkTable()
keyboard1 = telebot.types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
keyboard1.row('/start', 'Регистрация', 'Обо мне')
keyboard2 = telebot.types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
keyboard2.row('Создать таблицу', 'Калькулятор', 'Удалить таблицу', 'Добавить расход')
keyboard3 = telebot.types.InlineKeyboardMarkup()
keyboard3.row(telebot.types.InlineKeyboardButton('Да', callback_data='yes'),
              telebot.types.InlineKeyboardButton('Нет', callback_data='no'))
keyboard4 = telebot.types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
number = 0


@bot.message_handler(commands=['start'])
def start_message(message):
    print('Bot start')
    bot.send_message(message.chat.id, 'Привет! Я твой личный помошник по учету твоих финансов!', reply_markup=keyboard1)


@bot.message_handler(regexp="Обо мне")
def about_me(message):
    bot.send_message(message.chat.id, "Я бот, который будет учитывать твои расходы каждый день", reply_markup=keyboard1)


@bot.message_handler(regexp="Создать таблицу")
def create_table(message):
    user_id = message.chat.id
    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = 'Категории'
    book.save(f"{user_id}.xlsx")
    book.close()
    sent = bot.send_message(message.chat.id, 'Напиши категории затрат через пробел')
    bot.register_next_step_handler(sent, add_categories)


def add_categories(message):
    book = openpyxl.Workbook()
    sheet = book.active
    mes = message.text
    mes = str(mes)
    user_id = message.from_user.id
    categories = mes.split(' ')
    for i in range(len(categories)):
        sheet[f'A{i+2}'] = categories[i]
    book.save(f"{user_id}.xlsx")
    book.close()
    bot.send_message(message.chat.id, 'Готово!', reply_markup=keyboard2)


@bot.message_handler(regexp="Добавить расход")
def get_categories(message):
    user_id = message.chat.id
    book = openpyxl.open(f"{user_id}.xlsx", read_only=True)
    sheet = book.active
    for row in range(2, sheet.max_row + 1):
        categories = sheet[row][0].value
        print(categories)
        bot.send_message(message.chat.id, categories)
    sent = bot.send_message(message.chat.id, "Напиши категорию в которую хочешь добать расходы")
    bot.register_next_step_handler(sent, take_costs)


def take_costs(message):
    global number
    mes = message.text
    mes = str(mes)
    user_id = message.chat.id
    book = openpyxl.open(f"{user_id}.xlsx", read_only=True)
    sheet = book.active
    print(mes)
    for row in range(2, sheet.max_row + 1):
        if mes == sheet[row][0].value:
            number = row
            print(number)
    sent = bot.send_message(message.chat.id, "Напиши сумму расхода")
    bot.register_next_step_handler(sent, add_costs)


def add_costs(message):
    global number
    #wt.get_date(message)
    i = 1
    user_id = message.from_user.id
    print(number)
    book = openpyxl.open(f"{user_id}.xlsx", read_only=False)
    sheet = book.active
    mes = message.text
    mes = str(mes)
    sheet[number][1].value = mes
    bot.send_message(message.chat.id, 'Готово!', reply_markup=keyboard2)
    book.save(f"{user_id}.xlsx")
    book.close()
    i += 1


bot.polling(none_stop=True, interval=0)
